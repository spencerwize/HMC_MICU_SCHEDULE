#!/usr/bin/env python3
"""
HMC MICU APP Shift Schedule Builder
Apr 13 – Jul 19, 2026 (PP8–PP14)
"""

import os
import sys
import json
import random
from datetime import date, timedelta
from collections import defaultdict
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
import xlsxwriter

# ─────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────

STAFF = ["Katie", "John", "Hayden", "Todd", "Caroline",
         "Isabel", "Kristin", "Mandie", "Maureen", "Radha"]

STAFF_COL_MAP = {          # Column index in Time_Off xlsx (1-based)
    "Todd": 4, "Mandie": 5, "Isabel": 7, "Radha": 8,
    "Maureen": 9, "Katie": 10, "Hayden": 11, "Caroline": 12,
    "John": 13, "Kristin": 14,
}

SCHEDULE_START = date(2026, 4, 13)
SCHEDULE_END   = date(2026, 7, 19)

PAY_PERIODS = [
    ("PP8",  date(2026, 4, 13), date(2026, 4, 26)),
    ("PP9",  date(2026, 4, 27), date(2026, 5, 10)),
    ("PP10", date(2026, 5, 11), date(2026, 5, 24)),
    ("PP11", date(2026, 5, 25), date(2026, 6,  7)),
    ("PP12", date(2026, 6,  8), date(2026, 6, 21)),
    ("PP13", date(2026, 6, 22), date(2026, 7,  5)),
    ("PP14", date(2026, 7,  6), date(2026, 7, 19)),
]

# Holidays: date -> {slot: person}
HOLIDAYS = {
    date(2026, 5, 25): {"APP1": "Hayden", "APP2": "Todd",    "Roaming": "Radha",    "Night": "Isabel"},
    date(2026, 6, 19): {"APP1": "Mandie", "APP2": "Caroline","Roaming": "Radha",    "Night": "Isabel"},
    date(2026, 7,  4): {"APP1": "Kristin","APP2": "John",    "Roaming": "Caroline", "Night": "Mandie"},
}

# PP13 conference absences (Jun 22 – Jul 5)
PP13_CONFERENCE = {
    "John":   (date(2026, 6, 22), date(2026, 7, 5)),
    "Todd":   (date(2026, 6, 22), date(2026, 7, 5)),
    "Mandie": (date(2026, 6, 22), date(2026, 7, 5)),
    "Maureen":(date(2026, 6, 22), date(2026, 7, 5)),
}

SLOTS = ["APP1", "APP2", "Roaming", "Night"]
DAY_SLOTS = ["APP1", "APP2", "Roaming"]

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def all_dates():
    d = SCHEDULE_START
    while d <= SCHEDULE_END:
        yield d
        d += timedelta(1)

def get_pp(d):
    for name, start, end in PAY_PERIODS:
        if start <= d <= end:
            return name
    return None

def get_pp_dates(pp_name):
    for name, start, end in PAY_PERIODS:
        if name == pp_name:
            return [start + timedelta(i) for i in range((end - start).days + 1)]
    return []

def is_weekend(d):
    return d.weekday() >= 5  # Sat=5, Sun=6

# ─────────────────────────────────────────────
# PARSE TIME OFF FILE
# ─────────────────────────────────────────────

VAC_KEYWORDS = {"vac", "hawaii", "galapagos", "trip", "vacation", "travel"}

def color_matches(hex_color, target_prefix):
    """Check if color hex (AARRGGBB or RRGGBB) matches target prefix."""
    if not hex_color:
        return False
    c = hex_color.upper().lstrip("#")
    return c.startswith(target_prefix.upper())

def parse_time_off(xlsx_path):
    """
    Returns dict: person -> set of (date, type) where type in {off, vac, cme}
    off = plain off day (red fill, no vac keyword)
    vac = vacation day (red fill + vac keyword)
    cme = credited day (orange fill + yellow border)
    """
    result = {p: {} for p in STAFF}   # person -> date -> "off"/"vac"/"cme"

    if not os.path.exists(xlsx_path):
        print(f"WARNING: {xlsx_path} not found — proceeding with no time-off data.")
        return result

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_name = "2026 Full Year"
    if sheet_name not in wb.sheetnames:
        # Try first sheet
        ws = wb.active
        print(f"WARNING: Sheet '{sheet_name}' not found; using '{ws.title}'")
    else:
        ws = wb[sheet_name]

    # Find date rows: scan column A (or B) for date values
    # The xlsx likely has dates in rows, staff in columns
    # Row 1 = headers, column A = dates

    for row in ws.iter_rows():
        # Find the date in this row
        row_date = None
        for cell in row:
            if isinstance(cell.value, (date,)) or (
                hasattr(cell.value, 'date') and callable(cell.value.date)
            ):
                try:
                    row_date = cell.value.date() if hasattr(cell.value, 'date') else cell.value
                except:
                    pass
                break

        if row_date is None:
            continue
        if not (SCHEDULE_START <= row_date <= SCHEDULE_END):
            continue

        for person, col_idx in STAFF_COL_MAP.items():
            # col_idx is 1-based
            try:
                cell = row[col_idx - 1]
            except IndexError:
                continue

            fill = cell.fill
            font = cell.font
            border = cell.border
            val = str(cell.value).lower() if cell.value else ""

            # Detect fill color
            fg = None
            if fill and fill.fgColor:
                c = fill.fgColor
                if hasattr(c, 'rgb') and c.rgb:
                    fg = str(c.rgb).upper()
                elif hasattr(c, 'theme'):
                    fg = None

            # CME: orange fill (FFFF6D01) + yellow border on any side
            is_orange = fg and (fg == "FFFF6D01" or fg == "FF6D01")
            has_yellow_border = False
            if border:
                for side in [border.left, border.right, border.top, border.bottom]:
                    if side and side.color:
                        bc = side.color
                        bc_rgb = str(bc.rgb).upper() if hasattr(bc, 'rgb') and bc.rgb else ""
                        if bc_rgb in ("FFFFFF00", "FFFF00"):
                            has_yellow_border = True
                            break
            if is_orange and has_yellow_border:
                result[person][row_date] = "cme"
                continue

            # Red fill: FFF4CCCC or F4CCCC
            is_red = fg and (fg in ("FFF4CCCC", "F4CCCC") or fg.endswith("F4CCCC"))
            if is_red:
                # Check for vac keywords in cell value or nearby
                is_vac = any(kw in val for kw in VAC_KEYWORDS)
                result[person][row_date] = "vac" if is_vac else "off"

    return result


# ─────────────────────────────────────────────
# TARGET CALCULATION
# ─────────────────────────────────────────────

def compute_targets(time_off):
    """
    Returns per-person per-PP dict:
    {person: {pp_name: {avail, credited, target, sched_target, off_days, vac_days, cme_days}}}
    """
    targets = {}
    for person in STAFF:
        targets[person] = {}
        for pp_name, pp_start, pp_end in PAY_PERIODS:
            pp_dates = [pp_start + timedelta(i) for i in range((pp_end - pp_start).days + 1)]
            off_days = set()
            vac_days = set()
            cme_days = set()
            pdata = time_off.get(person, {})
            for d in pp_dates:
                t = pdata.get(d)
                if t == "off":
                    off_days.add(d)
                elif t == "vac":
                    vac_days.add(d)
                elif t == "cme":
                    cme_days.add(d)

            # PP13 conference absences → credited days
            if person in PP13_CONFERENCE:
                c_start, c_end = PP13_CONFERENCE[person]
                for d in pp_dates:
                    if c_start <= d <= c_end and d not in cme_days:
                        cme_days.add(d)

            credited = len(cme_days)
            all_off = off_days | vac_days | cme_days
            avail = len([d for d in pp_dates if d not in all_off])
            base = min(6, avail + credited)

            if avail >= 6:
                target = 6
            else:
                target = base

            sched_target = max(0, target - credited)

            targets[person][pp_name] = {
                "avail": avail,
                "credited": credited,
                "target": target,
                "sched_target": sched_target,
                "off_days": off_days,
                "vac_days": vac_days,
                "cme_days": cme_days,
                "pp_dates": pp_dates,
            }

    return targets


# ─────────────────────────────────────────────
# SCHEDULER
# ─────────────────────────────────────────────

class Scheduler:
    def __init__(self, time_off, targets):
        self.time_off = time_off
        self.targets = targets
        self.dates = list(all_dates())

        # schedule[date] = {slot: person or None}
        self.schedule = {}
        for d in self.dates:
            self.schedule[d] = {"APP1": None, "APP2": None, "Roaming": None, "Night": None}

        # Track per-person state
        self.person_shifts = {p: [] for p in STAFF}  # list of (date, slot)
        self.person_nights = {p: [] for p in STAFF}  # night dates
        self.pp_counts = {p: {pp: 0 for pp, *_ in PAY_PERIODS} for p in STAFF}
        self.pp_bumps = {p: {pp: False for pp, *_ in PAY_PERIODS} for p in STAFF}

    # ── Eligibility ──────────────────────────────────────────────────

    def is_blocked(self, person, d):
        """True if person cannot work at all on date d."""
        pdata = self.time_off.get(person, {})
        t = pdata.get(d)
        if t in ("off", "cme"):
            return True
        # Check PP13 conference
        if person in PP13_CONFERENCE:
            c_start, c_end = PP13_CONFERENCE[person]
            if c_start <= d <= c_end:
                return True
        return False

    def night_recovery_blocked(self, person, d, for_night=False):
        """
        After the LAST night of a streak ending on day N, blocked N+1 and N+2.
        for_night=True  -> allow streak continuation (last night == d-1 is OK)
        for_night=False -> strict (covers constraint 3 + recovery for day shifts)
        """
        past_nights = [nd for nd in self.person_nights[person] if nd < d]
        if not past_nights:
            return False
        last_night = max(past_nights)
        # Streak continuation: if we worked last night, keep going (nights only)
        if for_night and last_night == d - timedelta(1):
            return False
        return d <= last_night + timedelta(2)

    def had_night_on(self, person, d):
        return d in self.person_nights[person]

    def can_work_day(self, person, d):
        """Can person work a day shift on date d?"""
        if self.is_blocked(person, d):
            return False
        if self.night_recovery_blocked(person, d, for_night=False):
            return False
        # No day-to-night same day: check if assigned night on d
        # (can't work day AND night same day — handled at assignment time)
        # Check consecutive days limit (max 4)
        if self._consec_days_if_added(person, d) > 4:
            return False
        # Already assigned today?
        for slot in SLOTS:
            if self.schedule[d].get(slot) == person:
                return False
        return True

    def can_work_night(self, person, d):
        """Can person work night on date d?"""
        if self.is_blocked(person, d):
            return False
        if self.night_recovery_blocked(person, d, for_night=True):
            return False
        # No day shift same day (day-to-night constraint)
        for slot in DAY_SLOTS:
            if self.schedule[d].get(slot) == person:
                return False
        # Cannot work night if already pre-assigned to a day shift on D+1 or D+2
        # (covers holiday pre-seeds and the 2-day recovery window)
        for offset in (1, 2):
            check_d = d + timedelta(offset)
            if check_d in self.schedule:
                for slot in DAY_SLOTS:
                    if self.schedule[check_d].get(slot) == person:
                        return False
        # Max 3 consecutive nights
        consec = self._consec_nights_if_added(person, d)
        if consec > 3:
            return False
        # Max 4 consecutive working days
        if self._consec_days_if_added(person, d) > 4:
            return False
        # Already assigned tonight?
        if self.schedule[d].get("Night") == person:
            return False
        return True

    def _consec_nights_if_added(self, person, d):
        """Count consecutive nights ending on d (including d)."""
        count = 1
        prev = d - timedelta(1)
        while prev in self.person_nights[person]:
            count += 1
            prev -= timedelta(1)
        return count

    def _consec_days_if_added(self, person, d):
        """Count consecutive working days ending on d (including d)."""
        # Gather all worked dates
        worked = set(s for s, _ in self.person_shifts[person]) | set(self.person_nights[person])
        count = 1
        prev = d - timedelta(1)
        while prev in worked:
            count += 1
            prev -= timedelta(1)
        return count

    # ── Night eligibility score ──────────────────────────────────────

    def night_score(self, person, d):
        """Lower = prefer. Returns (priority, total_nights, pp_nights)."""
        nights = self.person_nights[person]
        # Streak continuation bonus (prefer continuing)
        is_continuation = (d - timedelta(1)) in nights
        pp = get_pp(d)
        pp_nights = sum(1 for nd in nights if get_pp(nd) == pp)
        total_nights = len(nights)
        # We want: continuation first, then fewest total nights, then fewest PP nights
        return (0 if is_continuation else 1, total_nights, pp_nights)

    # ── Day shift score ──────────────────────────────────────────────

    def day_score(self, person, d, slot):
        """Higher = prefer."""
        pp = get_pp(d)
        pp_info = self.targets[person][pp]
        days_left = len([x for x in pp_info["pp_dates"] if x >= d])
        slots_left = max(0, pp_info["sched_target"] - self.pp_counts[person][pp])
        urgency = slots_left / max(days_left, 1)

        # Cluster score
        worked = set(s for s, _ in self.person_shifts[person]) | set(self.person_nights[person])
        prev_worked = (d - timedelta(1)) in worked
        next_worked = (d + timedelta(1)) in worked
        if prev_worked and next_worked:
            cluster = 3
        elif prev_worked or next_worked:
            cluster = 2 if prev_worked else 1
        else:
            cluster = -3

        # Roaming balance
        roam_count = sum(1 for _, s in self.person_shifts[person] if s == "Roaming")
        total_shifts = len(self.person_shifts[person])
        roam_score = -roam_count / max(total_shifts, 1) if slot == "Roaming" else 0

        # Weekend fairness
        weekend_worked = sum(1 for s, _ in self.person_shifts[person] if is_weekend(s))
        weekend_ratio = weekend_worked / max(len(self.person_shifts[person]), 1)
        weekend_score = -weekend_ratio if is_weekend(d) else 0

        return urgency * 10 + cluster + roam_score + weekend_score

    # ── Pre-seed holidays ────────────────────────────────────────────

    def preseed_holidays(self):
        for d, assignments in HOLIDAYS.items():
            if d < SCHEDULE_START or d > SCHEDULE_END:
                continue
            for slot, person in assignments.items():
                self.schedule[d][slot] = person
                pp = get_pp(d)
                if pp and self.pp_counts[person].get(pp) is not None:
                    # Only count if not blocked (holiday overrides)
                    self.pp_counts[person][pp] += 1
                if slot == "Night":
                    self.person_nights[person].append(d)
                else:
                    self.person_shifts[person].append((d, slot))

    # ── Night scheduling (global, hardest-first) ─────────────────────

    def schedule_nights(self):
        dates = [d for d in self.dates if self.schedule[d]["Night"] is None]

        # Score each date by number of eligible candidates (hardest first)
        def date_difficulty(d):
            eligible = [p for p in STAFF if self.can_work_night(p, d)]
            return len(eligible)

        # Sort by difficulty (fewest eligible first), break ties by date
        dates_sorted = sorted(dates, key=lambda d: (date_difficulty(d), d))

        for d in dates_sorted:
            eligible = [p for p in STAFF if self.can_work_night(p, d)]
            pp = get_pp(d)
            if not eligible:
                # Force-fill with least-constrained person who is not blocked by off/cme
                eligible = [p for p in STAFF
                            if not self.is_blocked(p, d)
                            and not (d - timedelta(1)) in self.person_nights[p]  # avoid post-night same-day
                           ]
                if not eligible:
                    eligible = list(STAFF)

            # Check PP targets — prefer people under target
            under_target = [p for p in eligible
                            if pp and self.pp_counts[p].get(pp, 0) < self.targets[p][pp]["sched_target"]]
            pool = under_target if under_target else eligible

            # Score: continuation > fewest nights > fewest PP nights
            pool.sort(key=lambda p: self.night_score(p, d))
            chosen = pool[0]

            self.schedule[d]["Night"] = chosen
            self.person_nights[chosen].append(d)
            if pp:
                self.pp_counts[chosen][pp] += 1

    # ── Day shift scheduling (forward fill) ──────────────────────────

    def schedule_days(self):
        for d in self.dates:
            pp = get_pp(d)
            day_slots_needed = [s for s in DAY_SLOTS if self.schedule[d][s] is None]

            for slot in day_slots_needed:
                # Candidates: eligible, not night tonight, under target
                eligible = [p for p in STAFF if self.can_work_day(p, d)
                            and self.schedule[d]["Night"] != p]

                if not eligible:
                    continue

                under_target = [p for p in eligible
                                if pp and self.pp_counts[p].get(pp, 0) < self.targets[p][pp]["sched_target"]]
                pool = under_target if under_target else eligible

                # Sort by day score descending
                pool.sort(key=lambda p: -self.day_score(p, d, slot))
                chosen = pool[0]

                self.schedule[d][slot] = chosen
                self.person_shifts[chosen].append((d, slot))
                if pp:
                    self.pp_counts[chosen][pp] += 1

    # ── Cleanup pass ─────────────────────────────────────────────────

    def cleanup_pass(self):
        """Fill people who are under PP target."""
        for pp_name, pp_start, pp_end in PAY_PERIODS:
            pp_dates = [pp_start + timedelta(i) for i in range((pp_end - pp_start).days + 1)]
            for person in STAFF:
                pp_info = self.targets[person][pp_name]
                current = self.pp_counts[person][pp_name]
                needed = pp_info["sched_target"] - current
                if needed <= 0:
                    continue
                # Find open day slots on dates person hasn't worked
                for d in pp_dates:
                    if needed <= 0:
                        break
                    if not self.can_work_day(person, d):
                        continue
                    if self.schedule[d]["Night"] == person:
                        continue
                    for slot in DAY_SLOTS:
                        if self.schedule[d][slot] is None:
                            self.schedule[d][slot] = person
                            self.person_shifts[person].append((d, slot))
                            self.pp_counts[person][pp_name] += 1
                            needed -= 1
                            break

    # ── Force-fill APP1 ───────────────────────────────────────────────

    def force_fill_app1(self):
        """Ensure APP1 is always filled."""
        for d in self.dates:
            if self.schedule[d]["APP1"] is not None:
                continue
            pp = get_pp(d)
            # Find any available person
            eligible = [p for p in STAFF
                        if not self.is_blocked(p, d)
                        and self.schedule[d]["Night"] != p
                        and self.schedule[d]["APP2"] != p
                        and self.schedule[d]["Roaming"] != p
                        and not (self.had_night_on(p, d - timedelta(1)))]
            if not eligible:
                eligible = [p for p in STAFF if self.schedule[d]["Night"] != p]
            if not eligible:
                eligible = list(STAFF)

            # Prefer people least over target
            eligible.sort(key=lambda p: self.pp_counts[p].get(pp, 0))
            chosen = eligible[0]
            self.schedule[d]["APP1"] = chosen
            if (d, "APP1") not in self.person_shifts[chosen]:
                self.person_shifts[chosen].append((d, "APP1"))
            if pp:
                self.pp_counts[chosen][pp] += 1

    # ── Main run ─────────────────────────────────────────────────────

    def run(self):
        print("Pre-seeding holidays...")
        self.preseed_holidays()
        print("Scheduling nights (hardest-first)...")
        self.schedule_nights()
        print("Scheduling day shifts (forward fill)...")
        self.schedule_days()
        print("Cleanup pass...")
        self.cleanup_pass()
        print("Force-filling APP1...")
        self.force_fill_app1()
        print("Scheduling complete.")
        return self.schedule


# ─────────────────────────────────────────────
# EXCEL OUTPUT
# ─────────────────────────────────────────────

# Color constants (ARGB)
CLR_GREEN   = "FF92D050"   # day shift
CLR_BLUE    = "FFBDD7EE"   # night shift
CLR_PEACH   = "FFFFD966"   # vacation (yellow-ish peach)
CLR_PINK    = "FFFF99CC"   # PTO
CLR_ORANGE  = "FFFF6D01"   # CME
CLR_YELLOW  = "FFFFFF00"   # holiday
CLR_RED     = "FFFF0000"   # off (light red)
CLR_LIGHT_RED = "FFFFC7CE" # off
CLR_WHITE   = "FFFFFFFF"
CLR_HEADER  = "FF203864"   # dark blue header
CLR_HEADER2 = "FF2E75B6"   # medium blue
CLR_GRAY    = "FFD9D9D9"
CLR_WEEKEND = "FFF2F2F2"

def make_fill(argb):
    return PatternFill("solid", fgColor=argb)

def make_font(bold=False, color="FF000000", size=10):
    return Font(bold=bold, color=color, size=size)

def make_border(style="thin", color="FF000000"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def make_dashed_border(color="FFFF6D01"):
    s = Side(style="dashed", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def cell_value_and_color(person, d, slot, schedule, time_off, targets):
    """Determine what to show in a staff member's cell."""
    pp = get_pp(d)
    pdata = time_off.get(person, {})
    t = pdata.get(d)

    # Check conference/CME
    is_cme = t == "cme"
    if person in PP13_CONFERENCE:
        c_start, c_end = PP13_CONFERENCE[person]
        if c_start <= d <= c_end:
            is_cme = True

    if is_cme:
        return "CME", CLR_ORANGE, True  # value, color, is_cme

    # Check if working
    day_sched = schedule.get(d, {})
    for s in SLOTS:
        if day_sched.get(s) == person:
            role = "Night" if s == "Night" else ("APP1" if s == "APP1" else ("APP2" if s == "APP2" else "Roam"))
            color = CLR_BLUE if s == "Night" else CLR_GREEN
            # Holiday override
            if d in HOLIDAYS and HOLIDAYS[d].get(s) == person:
                color = CLR_YELLOW
            return role, color, False

    # Not working — determine why
    if t == "vac":
        # Check if PTO should be charged
        if pp:
            pp_info = targets[person][pp]
            if pp_info["avail"] < 6:
                return "PTO", CLR_PINK, False
        return "VAC", CLR_PEACH, False

    if t == "off":
        return "OFF", CLR_LIGHT_RED, False

    # Check if off due to conference (not cme-tagged)
    return "", CLR_WHITE, False


def build_excel(schedule, time_off, targets, output_path):
    wb = openpyxl.Workbook()

    # ── Sheet 3: Schedule (data source) ────────────────────────────
    ws_sched = wb.active
    ws_sched.title = "Schedule"

    headers = ["Date", "Day", "PP", "Shift"] + STAFF
    for col, h in enumerate(headers, 1):
        cell = ws_sched.cell(row=1, column=col, value=h)
        cell.fill = make_fill(CLR_HEADER)
        cell.font = make_font(bold=True, color="FFFFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = make_border()

    row_num = 2
    date_row_map = {}  # date -> (day_row, night_row)

    all_date_list = list(all_dates())

    for d in all_date_list:
        day_row = row_num
        night_row = row_num + 1
        date_row_map[d] = (day_row, night_row)

        pp = get_pp(d)
        day_name = d.strftime("%a")
        date_str = d.strftime("%m/%d/%Y")

        # Day row
        ws_sched.cell(row=day_row, column=1, value=date_str)
        ws_sched.cell(row=day_row, column=2, value=day_name)
        ws_sched.cell(row=day_row, column=3, value=pp)
        ws_sched.cell(row=day_row, column=4, value="Day 6:30A–6:30P")

        # Night row
        ws_sched.cell(row=night_row, column=1, value=date_str)
        ws_sched.cell(row=night_row, column=2, value=day_name)
        ws_sched.cell(row=night_row, column=3, value=pp)
        ws_sched.cell(row=night_row, column=4, value="Night 6:30P–6:30A")

        # Staff cells
        for col, person in enumerate(STAFF, 5):
            day_sched = schedule.get(d, {})

            # Day row: show day slot assignment
            day_role = None
            for slot in DAY_SLOTS:
                if day_sched.get(slot) == person:
                    day_role = slot
                    break

            # Night row: show night assignment
            night_role = "Night" if day_sched.get("Night") == person else None

            # Determine cell value and color for day row
            pdata = time_off.get(person, {})
            t = pdata.get(d)
            is_cme = t == "cme"
            if person in PP13_CONFERENCE:
                c_start, c_end = PP13_CONFERENCE[person]
                if c_start <= d <= c_end:
                    is_cme = True

            # Day row cell
            dcell = ws_sched.cell(row=day_row, column=col)
            ncell = ws_sched.cell(row=night_row, column=col)

            if is_cme:
                dcell.value = "CME"
                dcell.fill = make_fill(CLR_ORANGE)
                # CME border: orange + yellow
                yb = Side(style="medium", color="FFFFFF00")
                ob = Side(style="medium", color=CLR_ORANGE)
                dcell.border = Border(left=yb, right=yb, top=yb, bottom=yb)
                ncell.value = "CME"
                ncell.fill = make_fill(CLR_ORANGE)
                ncell.border = Border(left=yb, right=yb, top=yb, bottom=yb)
            elif day_role:
                abbr = "APP1" if day_role == "APP1" else ("APP2" if day_role == "APP2" else "Roam")
                dcell.value = abbr
                clr = CLR_YELLOW if d in HOLIDAYS else CLR_GREEN
                dcell.fill = make_fill(clr)
                dcell.border = make_border("thin", "FF000000")
                ncell.value = ""
                ncell.fill = make_fill(CLR_WHITE)
            elif t == "vac":
                pp_info = targets[person][pp] if pp else None
                if pp_info and pp_info["avail"] < 6:
                    dcell.value = "PTO"
                    dcell.fill = make_fill(CLR_PINK)
                else:
                    dcell.value = "VAC"
                    dcell.fill = make_fill(CLR_PEACH)
                dcell.border = make_border()
                ncell.value = ""
                ncell.fill = make_fill(CLR_WHITE)
            elif t == "off":
                dcell.value = "OFF"
                dcell.fill = make_fill(CLR_LIGHT_RED)
                dcell.border = make_border()
                ncell.value = ""
                ncell.fill = make_fill(CLR_WHITE)
            else:
                # Check soft constraint: day before someone's night
                next_night = schedule.get(d + timedelta(1), {}).get("Night")
                dcell.value = ""
                dcell.fill = make_fill(CLR_WEEKEND if is_weekend(d) else CLR_WHITE)

            # Night row cell
            if not is_cme:
                if night_role:
                    ncell.value = "Night"
                    clr = CLR_YELLOW if d in HOLIDAYS else CLR_BLUE
                    ncell.fill = make_fill(clr)
                    ncell.border = make_border("thin", "FF000000")
                elif not ncell.value:
                    ncell.fill = make_fill(CLR_WEEKEND if is_weekend(d) else CLR_WHITE)

            # Day-before-night soft flag: orange dashed border
            if not is_cme and not day_role and not t:
                # Check if person works night tomorrow
                tomorrow_night = schedule.get(d + timedelta(1), {}).get("Night")
                if tomorrow_night == person:
                    dcell.border = make_dashed_border("FFFF6D01")

            # Style formatting
            for cell in [dcell, ncell]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = make_font(size=9)

        # Formatting for fixed columns
        for rw in [day_row, night_row]:
            for col in range(1, 5):
                c = ws_sched.cell(row=rw, column=col)
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.font = make_font(size=9)
                bg = CLR_WEEKEND if is_weekend(d) else CLR_WHITE
                if col <= 2:
                    c.fill = make_fill(bg)
                elif col == 3:
                    c.fill = make_fill(CLR_GRAY)
                elif col == 4:
                    c.fill = make_fill("FFD6E4BC" if "Day" in (c.value or "") else "FFBDD7EE")

        row_num += 2

    # Column widths for Schedule sheet
    ws_sched.column_dimensions["A"].width = 12
    ws_sched.column_dimensions["B"].width = 6
    ws_sched.column_dimensions["C"].width = 6
    ws_sched.column_dimensions["D"].width = 18
    for col in range(5, 5 + len(STAFF)):
        ws_sched.column_dimensions[get_column_letter(col)].width = 10

    ws_sched.freeze_panes = "E2"

    # ── Sheet 2: Summary ────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    ws_sum.sheet_view.showGridLines = False

    # Title
    ws_sum.merge_cells("A1:L1")
    tc = ws_sum["A1"]
    tc.value = "HMC MICU APP Schedule — Apr 13 – Jul 19, 2026"
    tc.fill = make_fill(CLR_HEADER)
    tc.font = make_font(bold=True, color="FFFFFFFF", size=14)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 30

    # PP summary table header
    sum_headers = ["Person"] + [pp for pp, *_ in PAY_PERIODS] + ["Total Shifts", "Total Nights", "Total Roam"]
    for col, h in enumerate(sum_headers, 1):
        c = ws_sum.cell(row=3, column=col, value=h)
        c.fill = make_fill(CLR_HEADER2)
        c.font = make_font(bold=True, color="FFFFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = make_border()

    for row, person in enumerate(STAFF, 4):
        ws_sum.cell(row=row, column=1, value=person).font = make_font(bold=True)
        for col, (pp_name, *_) in enumerate(PAY_PERIODS, 2):
            count = sum(1 for d in all_date_list
                        for slot in SLOTS
                        if schedule[d].get(slot) == person and get_pp(d) == pp_name)
            cme = len(targets[person][pp_name]["cme_days"])
            c = ws_sum.cell(row=row, column=col, value=f"{count} / {targets[person][pp_name]['sched_target']} (CME:{cme})")
            c.alignment = Alignment(horizontal="center")
            c.border = make_border("thin", CLR_GRAY)
            if count < targets[person][pp_name]["sched_target"]:
                c.fill = make_fill("FFFFCCCC")
            elif count > targets[person][pp_name]["sched_target"]:
                c.fill = make_fill("FFCCFFCC")

        total_shifts = sum(1 for d in all_date_list for s in DAY_SLOTS if schedule[d].get(s) == person)
        total_nights = sum(1 for d in all_date_list if schedule[d].get("Night") == person)
        total_roam = sum(1 for d in all_date_list if schedule[d].get("Roaming") == person)
        ws_sum.cell(row=row, column=len(PAY_PERIODS) + 2, value=total_shifts).alignment = Alignment(horizontal="center")
        ws_sum.cell(row=row, column=len(PAY_PERIODS) + 3, value=total_nights).alignment = Alignment(horizontal="center")
        ws_sum.cell(row=row, column=len(PAY_PERIODS) + 4, value=total_roam).alignment = Alignment(horizontal="center")
        for col in range(1, len(sum_headers) + 1):
            ws_sum.cell(row=row, column=col).border = make_border("thin", CLR_GRAY)

    # PP13 note
    note_row = 4 + len(STAFF) + 2
    ws_sum.merge_cells(f"A{note_row}:L{note_row}")
    nc = ws_sum[f"A{note_row}"]
    nc.value = ("NOTE PP13 (Jun 22–Jul 5): John/Todd/Mandie/Maureen at conferences. "
                "Only 6 active staff × 6 shifts = 36 available vs 56 slots — 20 slots structurally empty.")
    nc.fill = make_fill("FFFFF2CC")
    nc.font = make_font(size=9, bold=True)
    nc.alignment = Alignment(wrap_text=True)

    # Metrics rows
    metrics_row = note_row + 2
    ws_sum.cell(row=metrics_row, column=1, value="Night Distribution").font = make_font(bold=True)
    for col, person in enumerate(STAFF, 2):
        total_nights = sum(1 for d in all_date_list if schedule[d].get("Night") == person)
        ws_sum.cell(row=metrics_row, column=col, value=f"{person}: {total_nights}")

    ws_sum.column_dimensions["A"].width = 12
    for col in range(2, len(sum_headers) + 1):
        ws_sum.column_dimensions[get_column_letter(col)].width = 18

    # ── Sheet 1: Calendar ────────────────────────────────────────────
    ws_cal = wb.create_sheet("Calendar", 0)
    ws_cal.sheet_view.showGridLines = False

    # Dropdown in C2 for staff selection (data validation)
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(STAFF) + '"',
        allow_blank=True,
        showDropDown=False
    )
    dv.sqref = "C2"
    ws_cal.add_data_validation(dv)

    ws_cal.cell(row=1, column=1, value="HMC MICU APP Schedule").font = make_font(bold=True, size=14)
    ws_cal.cell(row=2, column=1, value="Select Staff:").font = make_font(bold=True)
    ws_cal.cell(row=2, column=3, value=STAFF[0])  # default

    # Build calendar months: Apr, May, Jun, Jul
    months = [
        (2026, 4, "April 2026"),
        (2026, 5, "May 2026"),
        (2026, 6, "June 2026"),
        (2026, 7, "July 2026"),
    ]

    cal_row = 4
    DOW = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]

    # Build a lookup: date -> role for each person (static lookup from schedule)
    # We'll write a per-person static table approach since xlsx formulas referring
    # to a dropdown are complex. Instead we embed VBA-free formula lookup using
    # MATCH/INDEX on Schedule sheet.
    # For simplicity we write the values directly based on C2 selection using
    # a simple lookup formula.

    # We create a named table on Schedule sheet for VLOOKUP
    # Actually, for simplicity in pure openpyxl (no xlsxwriter VBA),
    # write static per-person data for currently selected person.
    # We'll write formulas referencing a lookup table.

    # Build a per-date lookup table on a helper sheet
    ws_lookup = wb.create_sheet("_Lookup")
    ws_lookup.sheet_view.showGridLines = False

    # Header: Date, Person1, Person2, ...
    ws_lookup.cell(row=1, column=1, value="Date")
    for ci, p in enumerate(STAFF, 2):
        ws_lookup.cell(row=1, column=ci, value=p)

    lookup_date_rows = {}
    for ri, d in enumerate(all_date_list, 2):
        ws_lookup.cell(row=ri, column=1, value=d.strftime("%m/%d/%Y"))
        lookup_date_rows[d] = ri
        for ci, person in enumerate(STAFF, 2):
            # Determine role
            day_sched = schedule.get(d, {})
            role = ""
            for slot in SLOTS:
                if day_sched.get(slot) == person:
                    if slot == "Night":
                        role = "Night"
                    elif slot == "APP1":
                        role = "APP1"
                    elif slot == "APP2":
                        role = "APP2"
                    else:
                        role = "Roam"
                    break
            if not role:
                pdata = time_off.get(person, {})
                t = pdata.get(d)
                is_cme = t == "cme"
                if person in PP13_CONFERENCE:
                    c_start, c_end = PP13_CONFERENCE[person]
                    if c_start <= d <= c_end:
                        is_cme = True
                if is_cme:
                    role = "CME"
                elif t == "vac":
                    pp = get_pp(d)
                    pp_info = targets[person][pp] if pp else None
                    role = "PTO" if (pp_info and pp_info["avail"] < 6) else "VAC"
                elif t == "off":
                    role = "OFF"
            ws_lookup.cell(row=ri, column=ci, value=role)

    # Calendar layout
    for year, month, month_label in months:
        # Month header
        ws_cal.merge_cells(f"A{cal_row}:G{cal_row}")
        mh = ws_cal[f"A{cal_row}"]
        mh.value = month_label
        mh.fill = make_fill(CLR_HEADER)
        mh.font = make_font(bold=True, color="FFFFFFFF", size=12)
        mh.alignment = Alignment(horizontal="center", vertical="center")
        ws_cal.row_dimensions[cal_row].height = 20
        cal_row += 1

        # Day-of-week header
        for ci, dow in enumerate(DOW):
            c = ws_cal.cell(row=cal_row, column=ci + 1, value=dow)
            c.fill = make_fill(CLR_HEADER2)
            c.font = make_font(bold=True, color="FFFFFFFF")
            c.alignment = Alignment(horizontal="center")
        cal_row += 1

        # Find first day of month
        first_day = date(year, month, 1)
        # Sunday=0, Monday=1... Python: Monday=0, Sunday=6
        # Adjust: Python weekday() Mon=0..Sun=6, we want Sun=0..Sat=6
        start_dow = (first_day.weekday() + 1) % 7  # 0=Sun

        # Find last day
        if month == 12:
            last_day = date(year + 1, 1, 1) - timedelta(1)
        else:
            last_day = date(year, month + 1, 1) - timedelta(1)

        week_row = cal_row
        col = start_dow + 1
        cur = first_day
        while cur <= last_day:
            if col > 7:
                col = 1
                week_row += 1

            cell_d = ws_cal.cell(row=week_row, column=col)
            # Show: date number + role for selected person
            # We use a formula to look up from _Lookup based on C2
            person_col_offset = None
            # We'll write date + value from _Lookup sheet using MATCH/INDEX
            # For static display, embed a formula
            lr = lookup_date_rows.get(cur)
            if lr:
                # =INDEX(_Lookup!$B$2:$K$200, MATCH($C$2, _Lookup!$B$1:$K$1, 0), lr-1)
                # Simpler: person col in _Lookup
                # Formula: date header + CHAR(10) + role
                staff_cols_range = "'_Lookup'!$B$1:$K$1"
                data_row = lr
                formula = (f'=TEXT({cur.strftime("%Y")*0 or ""}{cur.month}/{cur.day},"D")&" "&'
                           f'IFERROR(INDEX(\'_Lookup\'!$B${data_row}:$K${data_row},'
                           f'MATCH(C$2,{staff_cols_range},0)),"")')
                cell_d.value = formula
            else:
                cell_d.value = cur.day

            # Formatting
            if SCHEDULE_START <= cur <= SCHEDULE_END:
                cell_d.fill = make_fill(CLR_WEEKEND if is_weekend(cur) else CLR_WHITE)
            else:
                cell_d.fill = make_fill(CLR_GRAY)

            cell_d.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            cell_d.border = make_border("thin", CLR_GRAY)
            cell_d.font = make_font(size=9)

            cur += timedelta(1)
            col += 1

        cal_row = week_row + 2  # gap between months

    # Column widths for calendar
    for col in range(1, 8):
        ws_cal.column_dimensions[get_column_letter(col)].width = 14
    ws_cal.row_dimensions[2].height = 20

    # ── Save ────────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"Saved: {output_path}")


# ─────────────────────────────────────────────
# VALIDATION
# ─────────────────────────────────────────────

def validate(schedule, time_off, targets):
    errors = []
    warnings = []
    all_date_list = list(all_dates())

    # Constraint 1: APP1 always filled
    for d in all_date_list:
        if schedule[d]["APP1"] is None:
            errors.append(f"{d}: APP1 not filled")

    # Constraint 2: Night always filled
    for d in all_date_list:
        if schedule[d]["Night"] is None:
            errors.append(f"{d}: Night not filled")

    # Constraint 3 & 4: Night recovery
    for person in STAFF:
        nights = sorted(d for d in all_date_list if schedule[d].get("Night") == person)
        for nd in nights:
            for offset in [1, 2]:
                check_d = nd + timedelta(offset)
                if check_d in schedule:
                    for slot in DAY_SLOTS:
                        if schedule[check_d].get(slot) == person:
                            errors.append(f"{person}: working day on {check_d}, within {offset} day(s) of night on {nd}")
            # No day shift morning after night
            next_d = nd + timedelta(1)
            if next_d in schedule:
                for slot in DAY_SLOTS:
                    if schedule[next_d].get(slot) == person:
                        errors.append(f"{person}: day shift on {next_d} after night on {nd}")

    # Constraint 5: Max 3 consecutive nights
    for person in STAFF:
        nights = set(d for d in all_date_list if schedule[d].get("Night") == person)
        for d in all_date_list:
            if (d in nights and (d - timedelta(1)) in nights and
                    (d - timedelta(2)) in nights and (d - timedelta(3)) in nights):
                errors.append(f"{person}: 4+ consecutive nights ending {d}")

    # Constraint 7: No day-to-night same day
    for person in STAFF:
        for d in all_date_list:
            has_day = any(schedule[d].get(s) == person for s in DAY_SLOTS)
            has_night = schedule[d].get("Night") == person
            if has_day and has_night:
                errors.append(f"{person}: both day and night on {d}")

    # No double-booking
    for d in all_date_list:
        assigned = []
        for slot in SLOTS:
            p = schedule[d].get(slot)
            if p:
                assigned.append(p)
        if len(assigned) != len(set(assigned)):
            errors.append(f"{d}: double-booking detected: {assigned}")

    return errors, warnings


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    xlsx_input = os.path.join(os.path.dirname(__file__), "Time_Off_Requests.xlsx")
    output_path = os.path.join(os.path.dirname(__file__), "MICU_APP_Schedule_2026.xlsx")

    print("Parsing time-off data...")
    time_off = parse_time_off(xlsx_input)

    # Print parsed summary
    for person in STAFF:
        pdata = time_off.get(person, {})
        offs = sum(1 for v in pdata.values() if v == "off")
        vacs = sum(1 for v in pdata.values() if v == "vac")
        cmes = sum(1 for v in pdata.values() if v == "cme")
        print(f"  {person:10s}: off={offs}, vac={vacs}, cme={cmes}")

    print("\nComputing per-PP targets...")
    targets = compute_targets(time_off)

    for pp_name, *_ in PAY_PERIODS:
        print(f"\n  {pp_name}:")
        for person in STAFF:
            info = targets[person][pp_name]
            print(f"    {person:10s}: avail={info['avail']:2d}, credited={info['credited']}, "
                  f"target={info['target']}, sched_target={info['sched_target']}")

    print("\nRunning scheduler...")
    sched = Scheduler(time_off, targets)
    schedule = sched.run()

    print("\nValidating schedule...")
    errors, warnings = validate(schedule, time_off, targets)
    if errors:
        print(f"\nERRORS ({len(errors)}):")
        for e in errors[:20]:
            print(f"  ERROR: {e}")
        if len(errors) > 20:
            print(f"  ... and {len(errors)-20} more errors")
    else:
        print("  No hard constraint violations!")

    if warnings:
        for w in warnings[:10]:
            print(f"  WARN: {w}")

    print("\nBuilding Excel output...")
    build_excel(schedule, time_off, targets, output_path)

    # Print quick summary
    print("\nQuick shift counts:")
    all_date_list = list(all_dates())
    for person in STAFF:
        nights = sum(1 for d in all_date_list if schedule[d].get("Night") == person)
        days = sum(1 for d in all_date_list for s in DAY_SLOTS if schedule[d].get(s) == person)
        print(f"  {person:10s}: {days} day shifts, {nights} nights, total={days+nights}")


if __name__ == "__main__":
    main()
